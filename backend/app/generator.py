import json
import logging
import os
import re
from typing import Any, Dict, List, Optional, Tuple

from dotenv import load_dotenv
from langchain_core.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.config import OPENAI_MODEL, OUTPUT_DIR

load_dotenv()
logger = logging.getLogger(__name__)

REQUIRED_KEYS = [
    "Day",
    "Platform",
    "Content_Pillar",
    "Topic",
    "Hook",
    "Format",
    "Description",
    "CTA",
]

EXCEL_HEADERS = [
    ("Day", "Day"),
    ("Platform", "Platform"),
    ("Content_Pillar", "Content Pillar"),
    ("Topic", "Topic"),
    ("Hook", "Hook"),
    ("Format", "Format"),
    ("Description", "Description"),
    ("CTA", "CTA"),
]

PLATFORM_FORMATS = {
    "LinkedIn": ["thought leadership post", "carousel", "case-study post"],
    "Instagram": ["reel", "carousel", "story sequence"],
    "Twitter/X": ["thread", "short post", "poll"],
    "YouTube": ["short video", "long-form video", "community post"],
}

CONTENT_PILLARS = [
    "Education",
    "Authority",
    "Trust",
    "Lead Generation",
    "Product Story",
    "Community",
]

TOPIC_ANGLES = [
    "executive misconception",
    "before-and-after transformation",
    "buyer risk",
    "implementation roadmap",
    "team capability gap",
    "market timing",
    "operational bottleneck",
    "proof-of-value story",
    "decision checklist",
    "cost of delay",
    "customer education",
    "strategic advantage",
]

HOOK_PATTERNS = [
    "{audience} are not short on ambition. They are short on a clear {focus} execution path.",
    "The fastest way to waste budget on {focus} is to skip this one decision.",
    "Most teams approach {focus} from the tool side. The winners start with the workflow.",
    "If {audience} want better outcomes, this is the {focus} question to ask first.",
    "Here is the quiet gap that separates casual interest in {focus} from real adoption.",
    "A stronger {focus} plan starts before the first vendor conversation.",
    "The best signal that a team is ready for {focus} is not budget. It is clarity.",
    "Before {audience} invest in {focus}, they need to pressure-test this assumption.",
]

calendar_prompt = ChatPromptTemplate.from_template(
    """
You are a senior AI content strategist helping a user generate a sophisticated content calendar.

Create a content calendar using these inputs:
- Company details: {company_details}
- Weekly focus: {weekly_focus}
- Tone: {tone}
- Platforms: {platforms}
- Posts per day: {posts_per_day}
- Number of days: {number_of_days}
- Target audience: {target_audience}
- Call to action style: {call_to_action}

Requirements:
1. Generate exactly number_of_days * posts_per_day * number_of_platforms rows.
2. Use only the provided platforms.
3. Spread ideas intelligently across the full date sequence using labels like Day 1, Day 2, Day 3.
4. Make each row distinct and practical.
5. Return valid JSON only.
6. Each JSON object must contain these keys exactly:
   Day, Platform, Content_Pillar, Topic, Hook, Format, Description, CTA
7. Description must be useful enough that a content creator can execute the post immediately.
8. Hook must sound strong and platform-aware.
9. Format should fit the platform such as carousel, text post, short video, thread, reel, long-form video.
10. CTA should align with the provided call to action style.
11. No two rows may use the same Topic or Hook.
12. Avoid generic hooks. Make every hook specific to the platform, audience, and day.

Return a JSON array only. No markdown. No explanation.
"""
)


def extract_json(text: str) -> List[Dict[str, Any]]:
    if not text or not text.strip():
        raise ValueError("Empty model output")

    fenced = re.findall(r"```json\s*(.*?)\s*```", text, flags=re.DOTALL | re.IGNORECASE)
    if fenced:
        return json.loads(fenced[-1].strip())

    match = re.search(r"(\[.*\])", text, flags=re.DOTALL)
    if match:
        return json.loads(match.group(1))

    return json.loads(text)


def normalize_records(data: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    normalized = []
    for row in data if isinstance(data, list) else []:
        if not isinstance(row, dict):
            continue
        clean = {key: str(row.get(key, "")).strip() for key in REQUIRED_KEYS}
        normalized.append(clean)
    return normalized


def sanitize_file_name(file_name: str) -> str:
    safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "-", file_name or "")
    safe_name = re.sub(r"\s+", "-", safe_name).strip(" .-_")
    if not safe_name:
        safe_name = "content-calendar"
    if not safe_name.lower().endswith(".xlsx"):
        safe_name = f"{safe_name}.xlsx"
    return safe_name


def unique_file_name(file_name: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    base, extension = os.path.splitext(sanitize_file_name(file_name))
    candidate = f"{base}{extension}"
    counter = 2

    while os.path.exists(os.path.join(OUTPUT_DIR, candidate)):
        candidate = f"{base}-{counter}{extension}"
        counter += 1

    return candidate


def fallback_record(
    day: int,
    post_index: int,
    platform: str,
    company_details: str,
    weekly_focus: str,
    tone: str,
    call_to_action: str,
    target_audience: str,
) -> Dict[str, str]:
    formats = PLATFORM_FORMATS.get(platform, ["post"])
    seed = (day * 11) + (post_index * 7) + len(platform)
    content_pillar = CONTENT_PILLARS[seed % len(CONTENT_PILLARS)]
    post_format = formats[seed % len(formats)]
    angle = TOPIC_ANGLES[seed % len(TOPIC_ANGLES)]
    hook_pattern = HOOK_PATTERNS[seed % len(HOOK_PATTERNS)]
    focus = weekly_focus.strip()
    audience = target_audience.strip()
    platform_move = {
        "LinkedIn": "turn it into a boardroom-ready point of view",
        "Instagram": "make it visual, practical, and instantly understandable",
        "Twitter/X": "frame it as a sharp conversation starter",
        "YouTube": "expand it into a useful teaching moment",
    }.get(platform, "make it practical and platform-aware")

    return {
        "Day": f"Day {day}",
        "Platform": platform,
        "Content_Pillar": content_pillar,
        "Topic": f"{focus}: {angle} for {audience}",
        "Hook": hook_pattern.format(audience=audience, focus=focus),
        "Format": post_format,
        "Description": (
            f"Create a {post_format} for {platform} in a {tone} tone. Use the angle '{angle}' to connect "
            f"{company_details.strip()} with {focus}. Then {platform_move}, include one practical example, "
            f"and close with a clear next step."
        ),
        "CTA": call_to_action.strip(),
    }


def build_expected_records(
    company_details: str,
    weekly_focus: str,
    tone: str,
    platforms: List[str],
    posts_per_day: int,
    number_of_days: int,
    call_to_action: str,
    target_audience: str,
) -> List[Dict[str, str]]:
    records = []
    for day in range(1, number_of_days + 1):
        for post_index in range(1, posts_per_day + 1):
            for platform in platforms:
                records.append(
                    fallback_record(
                        day=day,
                        post_index=post_index,
                        platform=platform,
                        company_details=company_details,
                        weekly_focus=weekly_focus,
                        tone=tone,
                        call_to_action=call_to_action,
                        target_audience=target_audience,
                    )
                )
    return records


def generate_ai_records(
    company_details: str,
    weekly_focus: str,
    tone: str,
    platforms: List[str],
    posts_per_day: int,
    number_of_days: int,
    call_to_action: str,
    target_audience: str,
) -> Tuple[List[Dict[str, str]], Optional[str]]:
    try:
        llm = ChatOpenAI(model=OPENAI_MODEL, temperature=0.85, timeout=45, max_retries=1)
        response = (calendar_prompt | llm).invoke(
            {
                "company_details": company_details,
                "weekly_focus": weekly_focus,
                "tone": tone,
                "platforms": ", ".join(platforms),
                "posts_per_day": posts_per_day,
                "number_of_days": number_of_days,
                "target_audience": target_audience,
                "call_to_action": call_to_action,
            }
        )
        return normalize_records(extract_json(response.content)), None
    except Exception as error:
        logger.warning("AI calendar generation failed; using structured fallback. Error: %s", error)
        return [], error.__class__.__name__


def repair_records(ai_records: List[Dict[str, str]], expected_records: List[Dict[str, str]]) -> List[Dict[str, str]]:
    repaired = []

    for index, expected in enumerate(expected_records):
        ai_row = ai_records[index] if index < len(ai_records) else {}
        row = {key: ai_row.get(key) or expected[key] for key in REQUIRED_KEYS}
        row["Day"] = expected["Day"]
        row["Platform"] = expected["Platform"]
        repaired.append(row)

    return repaired


def generation_mode(ai_records: List[Dict[str, str]], expected_count: int, ai_error: Optional[str]) -> str:
    if ai_error or not ai_records:
        return "fallback"
    if len(ai_records) < expected_count:
        return "hybrid"
    return "ai"


def write_excel(records: List[Dict[str, str]], file_name: str) -> str:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Content Calendar"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin_side = Side(style="thin", color="B8C4D6")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for column_index, (_, label) in enumerate(EXCEL_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, record in enumerate(records, start=2):
        for column_index, (key, _) in enumerate(EXCEL_HEADERS, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=record.get(key, ""))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_widths = {
        "A": 12,
        "B": 16,
        "C": 18,
        "D": 34,
        "E": 48,
        "F": 20,
        "G": 64,
        "H": 34,
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    worksheet.row_dimensions[1].height = 28
    for row_index in range(2, len(records) + 2):
        worksheet.row_dimensions[row_index].height = 64

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    if records:
        table = Table(displayName="ContentCalendarTable", ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    file_path = os.path.join(OUTPUT_DIR, file_name)
    workbook.save(file_path)
    return file_path


def generate_content_calendar(
    company_details: str,
    weekly_focus: str,
    tone: str,
    platforms: List[str],
    posts_per_day: int,
    number_of_days: int,
    call_to_action: str,
    target_audience: str,
    output_file_name: str,
):
    expected_records = build_expected_records(
        company_details=company_details,
        weekly_focus=weekly_focus,
        tone=tone,
        platforms=platforms,
        posts_per_day=posts_per_day,
        number_of_days=number_of_days,
        call_to_action=call_to_action,
        target_audience=target_audience,
    )
    ai_records, ai_error = generate_ai_records(
        company_details=company_details,
        weekly_focus=weekly_focus,
        tone=tone,
        platforms=platforms,
        posts_per_day=posts_per_day,
        number_of_days=number_of_days,
        call_to_action=call_to_action,
        target_audience=target_audience,
    )
    records = repair_records(ai_records, expected_records)
    mode = generation_mode(ai_records, len(expected_records), ai_error)
    filename = unique_file_name(output_file_name)
    filepath = write_excel(records, filename)

    return {
        "records": records,
        "file_name": filename,
        "file_path": filepath,
        "generation_mode": mode,
        "warning": (
            "AI generation was unavailable, so ForgeFlow used its structured fallback engine."
            if mode == "fallback"
            else None
        ),
    }
